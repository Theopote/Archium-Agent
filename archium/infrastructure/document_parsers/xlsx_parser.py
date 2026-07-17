"""XLSX document parser."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from archium.infrastructure.document_parsers._utils import (
    _XLSX_SUFFIXES,
    build_parsed_document,
    normalize_whitespace,
    suffix_of,
)
from archium.infrastructure.document_parsers.base import ParsedDocument, ParsedPage

_MAX_ROWS_PER_SHEET = 200
_MAX_CELL_CHARS = 500


class XlsxParser:
    """Extract sheet-based table content from XLSX files."""

    def supports(self, file_path: Path) -> bool:
        return suffix_of(file_path) in _XLSX_SUFFIXES

    def parse(self, file_path: Path) -> ParsedDocument:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        pages: list[ParsedPage] = []
        metadata: dict[str, object] = {"sheet_names": workbook.sheetnames}

        for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]
            lines: list[str] = []
            row_count = 0

            for row in sheet.iter_rows(values_only=True):
                if row_count >= _MAX_ROWS_PER_SHEET:
                    lines.append("[sheet truncated]")
                    break
                values = [
                    self._stringify_cell(value)
                    for value in row
                    if value is not None and str(value).strip()
                ]
                if values:
                    lines.append(" | ".join(values))
                    row_count += 1

            text = normalize_whitespace("\n".join(lines))
            if text:
                pages.append(
                    ParsedPage(
                        page_number=sheet_index,
                        text=text,
                        section_title=sheet_name,
                        content_type="sheet",
                    )
                )

        workbook.close()
        return build_parsed_document(pages, metadata=metadata)

    def _stringify_cell(self, value: object) -> str:
        text = normalize_whitespace(str(value))
        if len(text) > _MAX_CELL_CHARS:
            return text[:_MAX_CELL_CHARS] + "..."
        return text
